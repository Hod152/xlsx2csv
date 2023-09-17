#! /bin/python3
# A tool to convert xlsx to csv
# Usage: ./xlsx2csv.py  [-i INPUT] [-s SHEET] [-o OUTDIR]

from openpyxl import load_workbook
from argparse import ArgumentParser
from csv import writer
from os import path

def xlsx2csv(wb, sheet_name, out_path):
    print("xlsx2csv.py: Writing to: ", out_path)
    with open(out_path, 'w', newline="") as f:
        c = writer(f)
        for r in wb[sheet_name].rows:
            c.writerow([cell.value for cell in r])

if __name__ == '__main__':
    # Parse arguments
    parser = ArgumentParser()
    parser.add_argument("-i", "--input", help="Path to xlsx file as an input")
    parser.add_argument("-s", "--sheet", help="Sheet name", required=False)
    parser.add_argument("-o", "--outdir", help="Output directory", required=False, default=".")
    args = parser.parse_args()

    WORKBOOK_PATH = args.input
    OUT_DIR = args.outdir

    wb = load_workbook(WORKBOOK_PATH)
    # Extract workbook name without xlsx ending
    wb_name = "".join(path.basename(WORKBOOK_PATH).split('.')[:-1])

    # Extract all sheets
    if args.sheet is None:
        for sheet in wb.sheetnames:
            try:
                OUT_PATH = path.join(OUT_DIR, wb_name + '_' + sheet + '.csv')
                xlsx2csv(wb, sheet, OUT_PATH)
            except Exception as e:
                print(e)
    # Extract particular sheet name
    else:
        sheet = args.sheet
        OUT_PATH = path.join(OUT_DIR, wb_name + '_' + sheet + '.csv')
        xlsx2csv(wb, sheet, OUT_PATH)